import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


HEADER_FILL = PatternFill(start_color="0A1F44", end_color="0A1F44", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
GREEN_FILL = PatternFill(start_color="2E8B57", end_color="2E8B57", fill_type="solid")
RED_FILL = PatternFill(start_color="C41E3A", end_color="C41E3A", fill_type="solid")
GOLD_FILL = PatternFill(start_color="D4A829", end_color="D4A829", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def style_header(ws, row=1, max_col=10, fill=None):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill or HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def auto_width(ws):
    for col in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 3, 50)


async def export_rapport_secteur(conn, secteur_code: str):
    """Export complet d'un secteur avec toutes les sources dynamiques."""
    is_industrie = secteur_code == "IND"
    secteur_nom = "Industrie" if is_industrie else "Commerce"

    wb = Workbook()

    # === Feuille 1: Indicateurs World Bank ===
    ws1 = wb.active
    ws1.title = f"World Bank - {secteur_nom}"

    if is_industrie:
        wb_codes = ['NV.IND.TOTL.ZS', 'NV.IND.TOTL.CD', 'NV.IND.MANF.ZS',
                     'SL.IND.EMPL.ZS', 'NE.EXP.GNFS.ZS', 'NE.GDI.FTOT.ZS',
                     'NY.GDP.MKTP.CD', 'NY.GDP.MKTP.KD.ZG', 'BX.KLT.DINV.CD.WD', 'SP.POP.TOTL']
    else:
        wb_codes = ['FP.CPI.TOTL.ZG', 'NE.EXP.GNFS.ZS', 'NE.IMP.GNFS.ZS',
                     'TG.VAL.TOTL.GD.ZS', 'NY.GDP.MKTP.CD', 'NY.GDP.MKTP.KD.ZG', 'SP.POP.TOTL']

    wb_rows = await conn.fetch("""
        SELECT indicateur_code, indicateur_nom, annee, valeur, unite
        FROM donnees_externes
        WHERE source_id = (SELECT id FROM sources_externes WHERE code = 'WORLDBANK')
          AND indicateur_code = ANY($1)
        ORDER BY indicateur_code, annee DESC
    """, wb_codes)

    headers1 = ["Code Indicateur", "Indicateur", "Annee", "Valeur", "Unite", "Source"]
    ws1.append(headers1)
    style_header(ws1, max_col=len(headers1))
    for r in wb_rows:
        ws1.append([r["indicateur_code"], r["indicateur_nom"], r["annee"],
                     float(r["valeur"]), r["unite"], "Banque Mondiale"])
    auto_width(ws1)

    # === Feuille 2: Trading Economics ===
    ws2 = wb.create_sheet(f"Trading Economics - {secteur_nom}")

    if is_industrie:
        te_codes = ['GDP', 'GDP_GROWTH', 'UNEMPLOYMENT', 'INTEREST_RATE']
    else:
        te_codes = ['EXPORTS', 'IMPORTS', 'TRADE_BALANCE', 'INFLATION', 'FOOD_INFLATION',
                     'GDP', 'GDP_GROWTH', 'UNEMPLOYMENT', 'PUBLIC_DEBT']

    te_rows = await conn.fetch("""
        SELECT indicateur_code, indicateur_nom, annee, valeur, unite
        FROM donnees_externes
        WHERE source_id = (SELECT id FROM sources_externes WHERE code = 'TRADINGECO')
          AND indicateur_code = ANY($1)
        ORDER BY indicateur_code
    """, te_codes)

    headers2 = ["Code", "Indicateur", "Annee", "Valeur", "Unite", "Source"]
    ws2.append(headers2)
    style_header(ws2, max_col=len(headers2), fill=GREEN_FILL if is_industrie else RED_FILL)
    for r in te_rows:
        ws2.append([r["indicateur_code"], r["indicateur_nom"], r["annee"],
                     float(r["valeur"]), r["unite"], "Trading Economics"])
    auto_width(ws2)

    # === Feuille 3: IPI / INS (industrie) ou Prix SIMPRIX (commerce) ===
    if is_industrie:
        ws3 = wb.create_sheet("IPI - INS Guinee")
        ipi_rows = await conn.fetch("""
            SELECT indicateur_code, indicateur_nom, annee, trimestre, valeur, unite, metadata
            FROM donnees_externes
            WHERE source_id = (SELECT id FROM sources_externes WHERE code = 'INS_GN')
            ORDER BY annee, trimestre
        """)
        headers3 = ["Indicateur", "Annee", "Trimestre", "Valeur (%)", "Unite", "Source"]
        ws3.append(headers3)
        style_header(ws3, max_col=len(headers3), fill=GREEN_FILL)
        for r in ipi_rows:
            ws3.append([r["indicateur_nom"], r["annee"], f"T{r['trimestre']}" if r["trimestre"] else "",
                         float(r["valeur"]), r["unite"], "INS Guinee"])
        auto_width(ws3)
    else:
        ws3 = wb.create_sheet("Prix - SIMPRIX")
        prix_rows = await conn.fetch("""
            SELECT nom_produit, categorie, prix_plafond, prix_marche, unite, date_releve
            FROM prix_produits ORDER BY nom_produit
        """)
        headers3 = ["Produit", "Categorie", "Prix Plafond", "Prix Marche", "Unite", "Date", "Source"]
        ws3.append(headers3)
        style_header(ws3, max_col=len(headers3), fill=RED_FILL)
        for r in prix_rows:
            ws3.append([r["nom_produit"], r["categorie"],
                         float(r["prix_plafond"]) if r["prix_plafond"] else None,
                         float(r["prix_marche"]) if r["prix_marche"] else None,
                         r["unite"], str(r["date_releve"]), "SIMPRIX"])
        auto_width(ws3)

    # === Feuille 4: Sous-secteurs ===
    ws4 = wb.create_sheet(f"Sous-secteurs {secteur_nom}")
    secteur_id = 1 if is_industrie else 2
    ss_rows = await conn.fetch(
        "SELECT code, nom, nom_en, code_isic, poids FROM sous_secteurs WHERE secteur_id = $1 ORDER BY ordre",
        secteur_id
    )
    headers4 = ["Code", "Nom", "Nom (EN)", "Code ISIC", "Poids (%)"]
    ws4.append(headers4)
    style_header(ws4, max_col=len(headers4), fill=GOLD_FILL)
    for r in ss_rows:
        ws4.append([r["code"], r["nom"], r["nom_en"], r["code_isic"],
                     float(r["poids"]) if r["poids"] else None])
    auto_width(ws4)

    # === Feuille 5: Contraintes ===
    ws5 = wb.create_sheet("Contraintes")
    ctr_rows = await conn.fetch("""
        SELECT c.nom, c.nom_en, cv.score, cv.nombre_repondants
        FROM contraintes_valeurs cv
        JOIN contraintes c ON c.id = cv.contrainte_id
        WHERE cv.secteur_id = $1
        ORDER BY c.ordre
    """, secteur_id)
    headers5 = ["Contrainte", "Contrainte (EN)", "Score (0-5)", "Repondants"]
    ws5.append(headers5)
    style_header(ws5, max_col=len(headers5))
    for r in ctr_rows:
        ws5.append([r["nom"], r["nom_en"], float(r["score"]), r["nombre_repondants"]])
    auto_width(ws5)

    # === Feuille 6: Indicateurs internes ===
    ws6 = wb.create_sheet(f"Indicateurs {secteur_nom}")
    ind_rows = await conn.fetch("""
        SELECT i.code, i.nom, i.nom_en, i.unite, i.source, i.periodicite, c.nom as categorie
        FROM indicateurs i
        JOIN categories c ON c.id = i.categorie_id
        JOIN secteurs s ON s.id = c.secteur_id
        WHERE s.code = $1 AND i.est_actif = true
        ORDER BY c.ordre, i.ordre
    """, secteur_code)
    headers6 = ["Code", "Indicateur", "Indicateur (EN)", "Categorie", "Unite", "Source", "Periodicite"]
    ws6.append(headers6)
    style_header(ws6, max_col=len(headers6))
    for r in ind_rows:
        ws6.append([r["code"], r["nom"], r["nom_en"], r["categorie"],
                     r["unite"], r["source"], r["periodicite"]])
    auto_width(ws6)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


async def export_indicateurs_excel(conn, secteur_code: str | None = None):
    query = """
        SELECT i.code, i.nom, i.unite, i.source, i.periodicite,
               c.nom as categorie, s.nom as secteur
        FROM indicateurs i
        JOIN categories c ON c.id = i.categorie_id
        JOIN secteurs s ON s.id = c.secteur_id
        WHERE i.est_actif = true
    """
    params = []
    if secteur_code:
        query += " AND s.code = $1"
        params.append(secteur_code)
    query += " ORDER BY s.nom, c.ordre, i.ordre"
    rows = await conn.fetch(query, *params)

    wb = Workbook()
    ws = wb.active
    ws.title = "Indicateurs"
    headers = ["Code", "Nom", "Secteur", "Categorie", "Unite", "Source", "Periodicite"]
    ws.append(headers)
    style_header(ws, max_col=len(headers))
    for row in rows:
        ws.append([row["code"], row["nom"], row["secteur"], row["categorie"],
                   row["unite"], row["source"], row["periodicite"]])
    auto_width(ws)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


async def export_valeurs_excel(conn, periode_id: int | None = None, secteur_code: str | None = None):
    # Use the new rapport export if secteur is specified
    if secteur_code:
        return await export_rapport_secteur(conn, secteur_code)

    query = """
        SELECT i.code, i.nom, i.unite, s.nom as secteur, c.nom as categorie,
               p.annee, p.trimestre, v.valeur, v.variation, v.tendance, v.statut,
               r.nom as region
        FROM valeurs v
        JOIN indicateurs i ON i.id = v.indicateur_id
        JOIN categories c ON c.id = i.categorie_id
        JOIN secteurs s ON s.id = c.secteur_id
        JOIN periodes p ON p.id = v.periode_id
        LEFT JOIN regions r ON r.id = v.region_id
        WHERE 1=1
    """
    params = []
    idx = 1
    if periode_id:
        query += f" AND v.periode_id = ${idx}"
        params.append(periode_id)
        idx += 1
    query += " ORDER BY p.annee, p.trimestre, s.nom, c.ordre, i.ordre"
    rows = await conn.fetch(query, *params)

    wb = Workbook()
    ws = wb.active
    ws.title = "Valeurs"
    headers = ["Code", "Indicateur", "Secteur", "Categorie", "Annee", "Trimestre",
               "Region", "Valeur", "Unite", "Variation (%)", "Tendance", "Statut"]
    ws.append(headers)
    style_header(ws, max_col=len(headers))
    for row in rows:
        ws.append([
            row["code"], row["nom"], row["secteur"], row["categorie"],
            row["annee"], row["trimestre"], row["region"] or "National",
            float(row["valeur"]), row["unite"],
            float(row["variation"]) if row["variation"] else None,
            row["tendance"], row["statut"]
        ])
    auto_width(ws)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
